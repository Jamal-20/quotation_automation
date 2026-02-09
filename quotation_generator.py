"""
JMK Auto-Q Genz
Automated Excel & PDF Quotation Generation Tool
Author: Jamaluddin
Date: 2026
"""

import sys
import os
import re
import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                            QHBoxLayout, QPushButton, QLabel, QLineEdit, 
                            QFileDialog, QMessageBox, QProgressBar, QTextEdit,
                            QGroupBox, QFrame)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QDragEnterEvent, QDropEvent, QCloseEvent

# Try to import win32com, but handle if Excel is not available
try:
    import win32com.client as win32
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


class QuotationWorker(QThread):
    progress = pyqtSignal(int)
    log = pyqtSignal(str)
    finished = pyqtSignal(bool, str)
    item_processing = pyqtSignal(str)  # For tracking current item
    
    def __init__(self, input_file, template_file, output_folder):
        super().__init__()
        self.input_file = input_file
        self.template_file = template_file
        self.output_folder = output_folder
        self._is_running = True
        self._is_cancelled = False
        
    def extract_number(self, text):
        if pd.isna(text):
            return 0.0
        text = str(text).strip()
        match = re.search(r'[\d.]+', text)
        if match:
            num_str = match.group()
            parts = num_str.split('.')
            if len(parts) > 2:
                num_str = parts[0] + '.' + ''.join(parts[1:])
            try:
                return float(num_str)
            except ValueError:
                return 0.0
        return 0.0
    
    def stop(self):
        self._is_running = False
        self._is_cancelled = True
        self.log.emit("⚠️ Cancellation requested... finishing current item...")
    
    def run(self):
        try:
            excel_path = os.path.join(self.output_folder, "Excel Output")
            pdf_path = os.path.join(self.output_folder, "PDF Output")
            
            os.makedirs(excel_path, exist_ok=True)
            os.makedirs(pdf_path, exist_ok=True)
            
            self.log.emit("📂 Loading input file...")
            if not os.path.exists(self.input_file):
                raise FileNotFoundError(f"Input file not found: {self.input_file}")
            
            if not os.path.exists(self.template_file):
                raise FileNotFoundError(f"Template file not found: {self.template_file}")
            
            df = pd.read_excel(self.input_file, header=None, dtype=str)
            
            while len(df.columns) < 12:
                df[len(df.columns)] = ""
            
            df.iloc[:, 11] = df.iloc[:, 10].apply(self.extract_number)
            
            last_row = df.shape[0]
            total = last_row - 1
            
            excel_app = None
            
            if HAS_EXCEL:
                try:
                    self.log.emit("🚀 Starting Excel application...")
                    excel_app = win32.Dispatch("Excel.Application")
                    excel_app.Visible = False
                    excel_app.DisplayAlerts = False
                    self.log.emit("✓ Excel started successfully")
                except Exception as e:
                    self.log.emit(f"⚠️ Warning: Could not start Excel - {str(e)}")
                    self.log.emit("📄 Will generate Excel files only (no PDF)")
            else:
                self.log.emit("⚠️ Warning: Excel COM not available")
                self.log.emit("📄 Will generate Excel files only (no PDF)")
            
            success_count = 0
            pdf_count = 0
            failed_items = []
            
            try:
                for i in range(1, last_row):
                    if not self._is_running:
                        self.log.emit("🛑 Process cancelled by user")
                        break
                    
                    sn_val = df.iloc[i, 1]
                    
                    if pd.isna(sn_val) or str(sn_val).strip() == "":
                        continue
                    
                    sn = str(sn_val).strip()
                    if sn == "":
                        sn = f"NoSN_{i+1}"
                    
                    self.item_processing.emit(sn)
                    percent = int((i / total) * 100)
                    self.progress.emit(percent)
                    
                    if not self._is_cancelled:
                        self.log.emit(f"⚙️ Processing {sn} ({i}/{total})...")
                    
                    try:
                        wb = load_workbook(self.template_file)
                        ws = wb.active
                        
                        ws['B11'] = df.iloc[i, 1]  # SN
                        ws['C11'] = df.iloc[i, 2]  # NUPCO
                        ws['D11'] = df.iloc[i, 3]  # Material Desc
                        ws['E11'] = df.iloc[i, 4]  # UOM
                        ws['F11'] = df.iloc[i, 5]  # Quantity
                        ws['G11'] = df.iloc[i, 6]  # Price
                        ws['H12'] = df.iloc[i, 7]  # VAT %
                        
                        excel_output = os.path.join(excel_path, f"Q{sn}.xlsx")
                        wb.save(excel_output)
                        wb.close()
                        success_count += 1
                        
                        if excel_app and not self._is_cancelled:
                            try:
                                pdf_output = os.path.join(pdf_path, f"Q{sn}.pdf")
                                xl_workbook = excel_app.Workbooks.Open(os.path.abspath(excel_output))
                                
                                if xl_workbook:
                                    xl_workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_output))
                                    xl_workbook.Close(SaveChanges=False)
                                    pdf_count += 1
                                else:
                                    self.log.emit(f"  ⚠️ Could not open {sn} in Excel for PDF")
                                    
                            except Exception as pdf_error:
                                self.log.emit(f"  ⚠️ PDF failed for {sn}: {str(pdf_error)}")
                                try:
                                    if 'xl_workbook' in locals() and xl_workbook:
                                        xl_workbook.Close(SaveChanges=False)
                                except:
                                    pass
                                    
                    except Exception as excel_error:
                        failed_items.append(sn)
                        self.log.emit(f"  ❌ Error processing {sn}: {str(excel_error)}")
                        continue
                
                if self._is_cancelled:
                    status_msg = f"⚠️ Cancelled! Generated {success_count} Excel files"
                else:
                    status_msg = f"✅ Done! Generated {success_count} Excel files"
                    if HAS_EXCEL:
                        status_msg += f" and {pdf_count} PDF files"
                
                self.progress.emit(100)
                self.log.emit(f"\n{status_msg}")
                self.log.emit(f"📁 Location: {self.output_folder}")
                
                if failed_items:
                    self.log.emit(f"\n⚠️ Failed items ({len(failed_items)}): {', '.join(failed_items[:5])}")
                    if len(failed_items) > 5:
                        self.log.emit(f"... and {len(failed_items) - 5} more")
                
                self.finished.emit(True, status_msg)
                
            finally:
                if excel_app:
                    try:
                        excel_app.Quit()
                    except:
                        pass
                    
        except Exception as e:
            self.log.emit(f"❌ Error: {str(e)}")
            self.finished.emit(False, str(e))


class DropArea(QFrame):
    fileDropped = pyqtSignal(str)
    
    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setMinimumHeight(80)
        self.setFrameStyle(QFrame.Shape.StyledPanel | QFrame.Shadow.Sunken)
        self.setStyleSheet("""
            DropArea {
                background-color: #f5f5f5;
                border: 2px dashed #999;
                border-radius: 8px;
            }
            DropArea:hover {
                background-color: #e8e8e8;
                border-color: #666;
            }
        """)
        
        layout = QVBoxLayout(self)
        self.label = QLabel(title)
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setFont(QFont("Segoe UI", 10))
        self.label.setStyleSheet("color: #666;")
        layout.addWidget(self.label)
        
        self.file_path = None
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            
    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.endswith(('.xlsx', '.xls')):
                self.file_path = file_path
                self.label.setText(f"✓ {os.path.basename(file_path)}")
                self.label.setStyleSheet("color: #4CAF50; font-weight: bold;")
                self.fileDropped.emit(file_path)
            else:
                QMessageBox.warning(self, "Invalid File", "Please drop an Excel file (.xlsx or .xls)")


class QuotationGenerator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AutoQ Genz by Jamal-JK v1.0")
        self.setMinimumSize(650, 600)
        self.worker = None
        self.setup_ui()
        
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        
        # Header
        header_layout = QVBoxLayout()
        title = QLabel("📝 AutoQ Genz")
        title.setFont(QFont("Segoe UI", 20, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #2c3e50;")
        header_layout.addWidget(title)
        
        subtitle = QLabel("Automated Excel & PDF Quotation Generation")
        subtitle.setFont(QFont("Segoe UI", 10))
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle.setStyleSheet("color: #7f8c8d;")
        header_layout.addWidget(subtitle)
        layout.addLayout(header_layout)
        
        # Status indicator
        self.status_label = QLabel()
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setFont(QFont("Segoe UI", 9))
        if HAS_EXCEL:
            self.status_label.setText("✓ System Optimized: Tender automation suite unlocked")
            self.status_label.setStyleSheet("color: #27ae60; padding: 5px;")
        else:
            self.status_label.setText("⚠️ Microsoft Excel not detected - Excel files only mode")
            self.status_label.setStyleSheet("color: #e67e22; padding: 5px;")
        layout.addWidget(self.status_label)
        
        # Input Section
        input_group = QGroupBox("📂 File Inputs (Drag & Drop or Browse)")
        input_layout = QVBoxLayout(input_group)
        input_layout.setSpacing(10)
        
        # Input File
        input_row = QHBoxLayout()
        self.input_drop = DropArea("Drop INPUT Excel File Here\n(or click Browse)")
        self.input_drop.fileDropped.connect(self.set_input_file)
        input_row.addWidget(self.input_drop, stretch=3)
        
        browse_input = QPushButton("Browse...")
        browse_input.setMinimumWidth(100)
        browse_input.clicked.connect(lambda: self.browse_file("input"))
        input_row.addWidget(browse_input, stretch=1)
        input_layout.addLayout(input_row)
        
        # Template File
        template_row = QHBoxLayout()
        self.template_drop = DropArea("Drop TEMPLATE Excel File Here\n(or click Browse)")
        self.template_drop.fileDropped.connect(self.set_template_file)
        template_row.addWidget(self.template_drop, stretch=3)
        
        browse_template = QPushButton("Browse...")
        browse_template.setMinimumWidth(100)
        browse_template.clicked.connect(lambda: self.browse_file("template"))
        template_row.addWidget(browse_template, stretch=1)
        input_layout.addLayout(template_row)
        
        layout.addWidget(input_group)
        
        # Output Section
        output_group = QGroupBox("💾 Output Location")
        output_layout = QHBoxLayout(output_group)
        
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("Select output folder...")
        self.output_path.setReadOnly(True)
        self.output_path.setStyleSheet("padding: 8px;")
        output_layout.addWidget(self.output_path)
        
        browse_output_btn = QPushButton("Browse...")
        browse_output_btn.setMinimumWidth(100)
        browse_output_btn.clicked.connect(self.browse_output)
        output_layout.addWidget(browse_output_btn)
        
        layout.addWidget(output_group)
        
        # Current item label
        self.current_item_label = QLabel("Ready_AutoQ Genz")
        self.current_item_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.current_item_label.setStyleSheet("color: #3498db; font-weight: bold;")
        layout.addWidget(self.current_item_label)
        
        # Progress Section
        progress_group = QGroupBox("⚡ Progress")
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 5px;
            }
        """)
        progress_layout.addWidget(self.progress_bar)
        
        # Log Area
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(180)
        self.log_area.setStyleSheet("""
            QTextEdit {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                font-family: 'Consolas', monospace;
                font-size: 9pt;
                padding: 5px;
            }
        """)
        progress_layout.addWidget(self.log_area)
        
        layout.addWidget(progress_group)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        
        self.generate_btn = QPushButton("🚀 GENERATE QUOTATIONS")
        self.generate_btn.setMinimumHeight(45)
        self.generate_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        self.generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #229954; }
            QPushButton:pressed { background-color: #1e8449; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.generate_btn.clicked.connect(self.start_generation)
        
        self.cancel_btn = QPushButton("⏹ CANCEL")
        self.cancel_btn.setMinimumHeight(45)
        self.cancel_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #c0392b; }
            QPushButton:pressed { background-color: #a93226; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.cancel_btn.clicked.connect(self.cancel_generation)
        
        btn_layout.addWidget(self.generate_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)
        
        # Status bar
        self.statusBar().showMessage("Ready")
        
    def set_input_file(self, path):
        self.input_drop.file_path = path
        
    def set_template_file(self, path):
        self.template_drop.file_path = path
        
    def browse_file(self, file_type):
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            f"Select {'Input' if file_type == 'input' else 'Template'} File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            if file_type == "input":
                self.input_drop.file_path = file_path
                self.input_drop.label.setText(f"✓ {os.path.basename(file_path)}")
                self.input_drop.label.setStyleSheet("color: #4CAF50; font-weight: bold;")
            else:
                self.template_drop.file_path = file_path
                self.template_drop.label.setText(f"✓ {os.path.basename(file_path)}")
                self.template_drop.label.setStyleSheet("color: #4CAF50; font-weight: bold;")
                
    def browse_output(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_path.setText(folder)
            
    def start_generation(self):
        input_file = getattr(self.input_drop, 'file_path', None)
        template_file = getattr(self.template_drop, 'file_path', None)
        output_folder = self.output_path.text()
        
        if not input_file:
            QMessageBox.warning(self, "Missing Input", "Please select an input Excel file.")
            return
        if not template_file:
            QMessageBox.warning(self, "Missing Template", "Please select a template Excel file.")
            return
        if not output_folder:
            QMessageBox.warning(self, "Missing Output", "Please select an output folder.")
            return
        
        self.generate_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.log_area.clear()
        self.current_item_label.setText("Starting...")
        
        self.worker = QuotationWorker(input_file, template_file, output_folder)
        self.worker.progress.connect(self.update_progress)
        self.worker.log.connect(self.add_log)
        self.worker.item_processing.connect(self.update_current_item)
        self.worker.finished.connect(self.generation_finished)
        self.worker.start()
        
    def update_progress(self, value):
        self.progress_bar.setValue(value)
        
    def update_current_item(self, item):
        self.current_item_label.setText(f"Processing: {item}")
        
    def add_log(self, message):
        self.log_area.append(message)
        
    def generation_finished(self, success, message):
        self.generate_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.current_item_label.setText("Ready")
        
        if success:
            QMessageBox.information(self, "Success", message)
            self.statusBar().showMessage("Generation completed")
        else:
            QMessageBox.critical(self, "Error", f"Generation failed:\n{message}")
            self.statusBar().showMessage("Generation failed")
            
    def cancel_generation(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.add_log("⏹ Cancellation requested...")
            self.cancel_btn.setEnabled(False)
            
    def closeEvent(self, event: QCloseEvent):
        """Handle window close event"""
        if self.worker and self.worker.isRunning():
            reply = QMessageBox.question(
                self, 
                "Confirm Exit",
                "⚠️ Generation is in progress!\n\nAre you sure you want to quit?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.worker.stop()
                self.worker.wait(2000)  # Wait max 2 seconds for graceful stop
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    # Set application-wide font
    font = QFont("Segoe UI", 9)
    app.setFont(font)
    
    window = QuotationGenerator()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()